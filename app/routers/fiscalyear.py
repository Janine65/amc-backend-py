"""``fiscalyear`` router (close-year + Bilanz/Erfolgsrechnung Excel)."""

from __future__ import annotations

import os
from datetime import UTC, date, datetime
from pathlib import Path
from typing import Annotated, Any

from fastapi import APIRouter, Depends, HTTPException, Query
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font
from openpyxl.worksheet.worksheet import Worksheet
from sqlalchemy import and_, func, or_, select
from sqlalchemy.ext.asyncio import AsyncSession

from app.core.config import get_config
from app.core.database import get_db
from app.core.logging import get_logger
from app.dependencies import CurrentUser
from app.models.account import Account
from app.models.budget import Budget
from app.models.fiscalyear import Fiscalyear
from app.models.journal import Journal
from app.schemas.fiscalyear import FiscalyearCreate, FiscalyearEntity, FiscalyearUpdate
from app.schemas.ret_data import RetData, RetDataFile, RetDataFilePayload
from app.utils.general import (
    I_FONT_SIZE_HEADER,
    I_FONT_SIZE_ROW,
    I_FONT_SIZE_TITEL,
    set_cell_value_format,
    thin_border,
)

logger = get_logger(__name__)

router = APIRouter(prefix="/fiscalyear", tags=["Fiscalyear"])

NUM_FMT = "#,##0.00;[Red]-#,##0.00"


@router.get("/getfiscalyearfk", response_model=RetData[list[dict]])
async def get_fk(_: CurrentUser, db: Annotated[AsyncSession, Depends(get_db)]) -> RetData[list[dict]]:
    rows = (await db.execute(select(Fiscalyear).order_by(Fiscalyear.year.desc()))).scalars().all()
    out: list[dict] = []
    for fy in rows:
        if fy.state == 1:
            css = "offen"
            name = f"{fy.name} - offen"
        elif fy.state == 2:
            css = "prov-closed"
            name = f"{fy.name} - prov. geschlossen"
        else:
            css = "closed"
            name = f"{fy.name} - geschlossen"
        out.append({**FiscalyearEntity.model_validate(fy).model_dump(), "name": name, "$css": css})
    return RetData(data=out, message="Fiscalyear FK")


@router.get("/getbyyear", response_model=RetData[FiscalyearEntity | None])
async def get_by_year(year: int, db: Annotated[AsyncSession, Depends(get_db)]) -> RetData[FiscalyearEntity | None]:
    row = await db.scalar(select(Fiscalyear).where(Fiscalyear.year == year))
    return RetData(
        data=FiscalyearEntity.model_validate(row) if row else None,
        message="Fiscalyear by year",
    )


@router.get("/closeyear", response_model=RetData[dict])
async def close_year(
    year: int = Query(...),
    state: int = Query(..., ge=2, le=3),
    _: CurrentUser = None,  # type: ignore[assignment]
    db: AsyncSession = Depends(get_db),
) -> RetData[dict]:
    """Close a fiscal year and create Saldovortrag bookings."""
    next_year = year + 1
    fy = await db.scalar(select(Fiscalyear).where(Fiscalyear.year == year))
    if fy is None:
        return RetData(
            data=None,
            type="error",
            message=f"Konnte Geschäftsjahr {year} nicht finden.",
        )

    next_fy = await db.scalar(select(Fiscalyear).where(Fiscalyear.year == next_year))
    cfg = get_config()
    now = datetime.now(UTC)
    if next_fy is None:
        next_fy = Fiscalyear(
            year=next_year,
            name=f"AMC-Buchhaltung {next_year}",
            state=1,
            createdAt=now,
            updatedAt=now,
        )
        db.add(next_fy)
        os.makedirs(f"{cfg.documents}{next_year}/receipt", exist_ok=True)

    # delete previous Saldovortrag rows
    next_journals = (
        (
            await db.execute(
                select(Journal).where(
                    and_(
                        Journal.year == next_year,
                        or_(Journal.from_account == 39, Journal.to_account == 39),
                    )
                )
            )
        )
        .scalars()
        .all()
    )
    for j in next_journals:
        await db.delete(j)

    igewinn = 0.0

    async def grouped(account_attr, level: int, year_: int, exclude39: bool = False):
        col = getattr(Journal, account_attr)
        join = Account.id == col
        stmt = (
            select(col, func.sum(Journal.amount).label("amount"))
            .join(Account, join)
            .where(and_(Journal.year == year_, Account.level == level))
            .group_by(col)
            .order_by(col.asc())
        )
        if exclude39:
            stmt = stmt.where(col != 39)
        rows = (await db.execute(stmt)).all()
        return [(r[0], float(r[1] or 0)) for r in rows]

    booking_date = date(next_year, 1, 1)
    journal_data: list[dict[str, Any]] = []

    # Aktiv (level 1): from_account
    for acc_id, amt in await grouped("from_account", 1, year):
        igewinn += amt
        journal_data.append(
            {
                "date": booking_date,
                "year": next_year,
                "memo": "Kontoeröffnung (Saldovortrag)",
                "amount": round(amt, 2),
                "from_account": acc_id,
                "to_account": 39,
            }
        )
    for acc_id, amt in await grouped("to_account", 1, year):
        igewinn -= amt
        existing = next((d for d in journal_data if d["from_account"] == acc_id), None)
        if existing:
            existing["amount"] = round(existing["amount"] - amt, 2)
        else:
            journal_data.append(
                {
                    "date": booking_date,
                    "year": next_year,
                    "memo": "Kontoeröffnung (Saldovortrag)",
                    "amount": round(-amt, 2),
                    "from_account": acc_id or 0,
                    "to_account": 39,
                }
            )
    # Passiv (level 2): from_account excluded 39
    for acc_id, amt in await grouped("from_account", 2, year, exclude39=True):
        igewinn -= amt
        journal_data.append(
            {
                "date": booking_date,
                "year": next_year,
                "memo": "Kontoeröffnung (Saldovortrag)",
                "amount": round(amt, 2),
                "from_account": 39,
                "to_account": acc_id,
            }
        )
    for acc_id, amt in await grouped("to_account", 2, year, exclude39=True):
        igewinn += amt
        existing = next((d for d in journal_data if d["to_account"] == acc_id and d["from_account"] == 39), None)
        if existing:
            existing["amount"] = round(existing["amount"] - amt, 2)
        else:
            journal_data.append(
                {
                    "date": booking_date,
                    "year": next_year,
                    "memo": "Kontoeröffnung (Saldovortrag)",
                    "amount": round(-amt, 2),
                    "from_account": 39,
                    "to_account": acc_id,
                }
            )

    for i, jd in enumerate(journal_data, start=1):
        db.add(Journal(**jd, journalno=i, createdAt=now, updatedAt=now))

    # Copy budget if next year's budget is empty
    nb_count = await db.scalar(select(func.count(Budget.id)).where(Budget.year == next_year))
    if not nb_count:
        prev = (await db.execute(select(Budget).where(Budget.year == year))).scalars().all()
        for b in prev:
            db.add(
                Budget(
                    account=b.account,
                    year=next_year,
                    memo=b.memo,
                    amount=b.amount,
                    createdAt=now,
                    updatedAt=now,
                )
            )

    fy.state = state
    fy.updatedAt = now
    await db.flush()
    igewinn = round(igewinn, 2)
    return RetData(
        data={"gewinn": igewinn},
        type="info",
        message=f"AMC-Buchhaltung {year} wurde erfolgreich beendet mit Gewinn/Verlust {igewinn:.2f}",
    )


# ---------------------------------------------------------------------------
# Bilanz / Erfolgsrechnung / Budget Excel export
# ---------------------------------------------------------------------------


def _font_t(bold: bool = False) -> Font:
    return Font(name="Tahoma", size=I_FONT_SIZE_TITEL, bold=bold)


def _font_r(bold: bool = False) -> Font:
    return Font(name="Tahoma", size=I_FONT_SIZE_ROW, bold=bold)


def _write_array(
    sheet: Worksheet,
    bilanz_list: list[dict],
    first_row: int,
    f_budget: bool = False,
    f_budget_vergleich: bool = False,
) -> dict[str, int]:
    """Hierarchical writer mirroring ``writeArray`` in fiscalyear.service.ts."""
    row = first_row
    cell_level = 0
    for element in bilanz_list:
        if element["level"] == element["order"]:
            row += 1
            cell_level = row
            set_cell_value_format(
                sheet,
                f"B{row}:C{row}",
                element["name"],
                border=True,
                merge=True,
                font=_font_t(bold=True),
            )
            for col in ("D", "E", "F"):
                set_cell_value_format(sheet, f"{col}{row}", "", border=True, font=_font_t(bold=True))
                sheet[f"{col}{row}"].number_format = NUM_FMT
            if f_budget:
                for col in ("G", "H"):
                    set_cell_value_format(sheet, f"{col}{row}", "", border=True, font=_font_t(bold=True))
                    sheet[f"{col}{row}"].number_format = NUM_FMT
        else:
            font = _font_r()
            set_cell_value_format(sheet, f"B{row}", element["order"], border=True, font=font)
            set_cell_value_format(sheet, f"C{row}", element["name"], border=True, font=font)
            set_cell_value_format(sheet, f"D{row}", element["amount"], border=True, font=font, align_h="right")
            set_cell_value_format(sheet, f"E{row}", element["amountVJ"], border=True, font=font, align_h="right")

            level = element["level"]
            if level in (2, 4):
                sheet[f"F{row}"] = f"=E{row}-D{row}"
            elif level in (1, 6):
                sheet[f"F{row}"] = f"=D{row}-E{row}"
            sheet[f"F{row}"].border = thin_border()
            sheet[f"F{row}"].font = font
            sheet[f"F{row}"].alignment = Alignment(horizontal="right")
            for col in ("D", "E", "F"):
                sheet[f"{col}{row}"].number_format = NUM_FMT

            if cell_level:
                sheet[f"D{cell_level}"] = f"=SUM(D{cell_level + 1}:D{row})"
                sheet[f"E{cell_level}"] = f"=SUM(E{cell_level + 1}:E{row})"
                sheet[f"F{cell_level}"] = f"=SUM(F{cell_level + 1}:F{row})"

            if f_budget:
                set_cell_value_format(sheet, f"G{row}", element["budget"], border=True, font=font, align_h="right")
                if level in (2, 4):
                    sheet[f"H{row}"] = f"=G{row}-D{row}"
                elif level in (1, 6):
                    sheet[f"H{row}"] = f"=D{row}-G{row}"
                sheet[f"H{row}"].border = thin_border()
                sheet[f"H{row}"].font = font
                sheet[f"H{row}"].alignment = Alignment(horizontal="right")
                if cell_level:
                    sheet[f"G{cell_level}"] = f"=SUM(G{cell_level + 1}:G{row})"
                    sheet[f"H{cell_level}"] = f"=SUM(H{cell_level + 1}:H{row})"
                for col in ("G", "H"):
                    sheet[f"{col}{row}"].number_format = NUM_FMT

            if f_budget_vergleich:
                set_cell_value_format(sheet, f"E{row}", element["budget"], border=True, font=font, align_h="right")
                sheet[f"F{row}"] = f"=E{row}-D{row}"
                set_cell_value_format(sheet, f"G{row}", element["budgetNJ"], border=True, font=font, align_h="right")
                sheet[f"H{row}"] = f"=G{row}-E{row}"
        row += 1
    return {"lastRow": row - 1, "total1": first_row + 1, "total2": cell_level}


@router.get("/writebilanz", response_model=RetDataFile)
async def write_bilanz(year: int, _: CurrentUser, db: Annotated[AsyncSession, Depends(get_db)]) -> RetDataFile:
    cfg = get_config()
    iv = year - 1
    inj = year + 1

    workbook = Workbook()
    workbook.remove(workbook.active)
    bsheet = workbook.create_sheet("Bilanz")
    esheet = workbook.create_sheet("Erfolgsrechnung")
    busheet = workbook.create_sheet("Budget")

    accounts = (await db.execute(select(Account).order_by(Account.level.asc(), Account.order.asc()))).scalars().all()
    budgets = (
        (
            await db.execute(
                select(Budget).where(Budget.year.in_([iv, year, inj])).order_by(Budget.year.asc(), Budget.account.asc())
            )
        )
        .scalars()
        .all()
    )

    bilanz: list[dict] = []
    for a in accounts:
        bilanz.append(
            {
                "id": a.id,
                "name": a.longname or f"{a.order} {a.name}",
                "level": a.level,
                "order": a.order,
                "status": a.status,
                "amount": 0.0,
                "amountVJ": 0.0,
                "amountNJ": 0.0,
                "budget": 0.0,
                "budgetVJ": 0.0,
                "budgetNJ": 0.0,
            }
        )
    for b in budgets:
        for rec in bilanz:
            if rec["id"] == b.account:
                if b.year == year:
                    rec["budget"] = float(b.amount or 0)
                elif b.year == iv:
                    rec["budgetVJ"] = float(b.amount or 0)
                elif b.year == inj:
                    rec["budgetNJ"] = float(b.amount or 0)
                break

    async def grouped(col_attr: str, year_: int) -> dict[int, float]:
        col = getattr(Journal, col_attr)
        rows = (
            await db.execute(
                select(col, func.sum(Journal.amount).label("amount")).where(Journal.year == year_).group_by(col)
            )
        ).all()
        return {r[0]: float(r[1] or 0) for r in rows if r[0] is not None}

    fa = await grouped("from_account", year)
    ta = await grouped("to_account", year)
    fv = await grouped("from_account", iv)
    tv = await grouped("to_account", iv)

    for rec in bilanz:
        rec["amount"] += fa.get(rec["id"], 0)
        rec["amountVJ"] += fv.get(rec["id"], 0)
        if rec["level"] in (1, 4):
            rec["amount"] -= ta.get(rec["id"], 0)
            rec["amountVJ"] -= tv.get(rec["id"], 0)
        elif rec["level"] in (2, 3):
            rec["amount"] = ta.get(rec["id"], 0) - rec["amount"]
            rec["amountVJ"] = tv.get(rec["id"], 0) - rec["amountVJ"]

    # ---------- Bilanz ----------
    set_cell_value_format(bsheet, "B1", f"Bilanz {year}", font=Font(name="Tahoma", size=I_FONT_SIZE_HEADER, bold=True))
    headers = [
        ("B3", "Konto"),
        ("C3", "Bezeichnung"),
        ("D3", f"Saldo {year}"),
        ("E3", f"Saldo {iv}"),
        ("F3", "Differenz"),
    ]
    for cell, txt in headers:
        set_cell_value_format(
            bsheet,
            cell,
            txt,
            border=True,
            font=_font_t(True),
            align_h="right" if cell.startswith(("D", "E", "F")) else "left",
        )
    bilanz_data = [b for b in bilanz if (b["status"] == 1 or b["amount"] or b["amountVJ"]) and (b["level"] or 0) < 3]
    total = _write_array(bsheet, bilanz_data, 4, f_budget=False)
    row = total["lastRow"] + 2
    set_cell_value_format(
        bsheet,
        f"B{row}:C{row}",
        "Gewinn / Verlust",
        border=True,
        merge=True,
        font=Font(name="Tahoma", size=I_FONT_SIZE_HEADER, bold=True),
    )
    bsheet[f"D{row}"] = f"=D{total['total1']}-D{total['total2']}"
    bsheet[f"E{row}"] = f"=E{total['total1']}-E{total['total2']}"
    bsheet[f"F{row}"] = f"=D{row}-E{row}"
    for col in ("D", "E", "F"):
        bsheet[f"{col}{row}"].number_format = NUM_FMT
        bsheet[f"{col}{row}"].font = _font_t(True)
        bsheet[f"{col}{row}"].border = thin_border()
    row += 2
    set_cell_value_format(
        bsheet,
        f"B{row}:C{row}",
        "Vermögen Ende Jahr",
        border=True,
        merge=True,
        font=Font(name="Tahoma", size=I_FONT_SIZE_HEADER, bold=True),
    )
    bsheet[f"D{row}"] = f"=D{total['lastRow']}+D{total['lastRow'] + 2}"
    bsheet[f"E{row}"] = f"=E{total['lastRow']}+E{total['lastRow'] + 2}"
    bsheet[f"F{row}"] = f"=D{row}-E{row}"
    for col in ("D", "E", "F"):
        bsheet[f"{col}{row}"].number_format = NUM_FMT
        bsheet[f"{col}{row}"].font = _font_t(True)
        bsheet[f"{col}{row}"].border = thin_border()
    for col, w in [("C", 32), ("D", 18), ("E", 18), ("F", 18)]:
        bsheet.column_dimensions[col].width = w

    # ---------- Erfolgsrechnung ----------
    set_cell_value_format(
        esheet,
        "B1",
        f"Erfolgsrechnung {year}",
        font=Font(name="Tahoma", size=I_FONT_SIZE_HEADER, bold=True),
    )
    e_headers = [
        ("B3", "Konto"),
        ("C3", "Bezeichnung"),
        ("D3", f"Saldo {year}"),
        ("E3", f"Saldo {iv}"),
        ("F3", "Differenz"),
        ("G3", f"Budget {year}"),
        ("H3", "Differenz"),
    ]
    for cell, txt in e_headers:
        set_cell_value_format(
            esheet,
            cell,
            txt,
            border=True,
            font=_font_t(True),
            align_h="right" if cell[0] in "DEFGH" else "left",
        )
    e_data = [
        b
        for b in bilanz
        if (b["status"] == 1 or b["amount"] or b["amountVJ"] or b["budget"] or b["budgetNJ"])
        and 2 < (b["level"] or 0) < 9
    ]
    total = _write_array(esheet, e_data, 4, f_budget=True)
    row = total["lastRow"] + 2
    set_cell_value_format(
        esheet,
        f"B{row}:C{row}",
        "Gewinn / Verlust",
        border=True,
        merge=True,
        font=Font(name="Tahoma", size=I_FONT_SIZE_HEADER, bold=True),
    )
    esheet[f"D{row}"] = f"=D{total['total2']}-D{total['total1']}"
    esheet[f"E{row}"] = f"=E{total['total2']}-E{total['total1']}"
    esheet[f"F{row}"] = f"=D{row}-E{row}"
    esheet[f"G{row}"] = f"=G{total['total2']}-G{total['total1']}"
    esheet[f"H{row}"] = f"=G{row}-D{row}"
    for col in ("D", "E", "F", "G", "H"):
        esheet[f"{col}{row}"].number_format = NUM_FMT
        esheet[f"{col}{row}"].font = _font_t(True)
        esheet[f"{col}{row}"].border = thin_border()
    for col, w in [("C", 32), ("D", 18), ("E", 18), ("F", 18), ("G", 18), ("H", 18)]:
        esheet.column_dimensions[col].width = w

    # ---------- Budgetvergleich ----------
    set_cell_value_format(busheet, "B1", f"Budget {inj}", font=Font(name="Tahoma", size=I_FONT_SIZE_HEADER, bold=True))
    b_headers = [
        ("B3", "Konto"),
        ("C3", "Bezeichnung"),
        ("D3", f"Saldo {year}"),
        ("E3", f"Budget {year}"),
        ("F3", "Differenz"),
        ("G3", f"Budget {inj}"),
        ("H3", "Differenz"),
    ]
    for cell, txt in b_headers:
        set_cell_value_format(
            busheet,
            cell,
            txt,
            border=True,
            font=_font_t(True),
            align_h="right" if cell[0] in "DEFGH" else "left",
        )
    total = _write_array(busheet, e_data, 4, f_budget=True, f_budget_vergleich=True)
    row = total["lastRow"] + 2
    set_cell_value_format(
        busheet,
        f"B{row}:C{row}",
        "Gewinn / Verlust",
        border=True,
        merge=True,
        font=Font(name="Tahoma", size=I_FONT_SIZE_HEADER, bold=True),
    )
    busheet[f"D{row}"] = f"=D{total['total2']}-D{total['total1']}"
    busheet[f"E{row}"] = f"=E{total['total2']}-E{total['total1']}"
    busheet[f"F{row}"] = f"=E{row}-D{row}"
    busheet[f"G{row}"] = f"=G{total['total2']}-G{total['total1']}"
    busheet[f"H{row}"] = f"=G{row}-E{row}"
    for col in ("D", "E", "F", "G", "H"):
        busheet[f"{col}{row}"].number_format = NUM_FMT
        busheet[f"{col}{row}"].font = _font_t(True)
        busheet[f"{col}{row}"].border = thin_border()
    for col, w in [("C", 32), ("D", 18), ("E", 18), ("F", 18), ("G", 18), ("H", 18)]:
        busheet.column_dimensions[col].width = w

    filename = f"Bilanz-{year}.xlsx"
    Path(cfg.exports).mkdir(parents=True, exist_ok=True)
    workbook.save(cfg.exports + filename)
    return RetDataFile(
        data=RetDataFilePayload(filename=filename),
        message=f"Bilanz {year} wurde erfolgreich erstellt",
    )


# ---------------------------------------------------------------------------
# CRUD
# ---------------------------------------------------------------------------


@router.post("", response_model=RetData[FiscalyearEntity], status_code=201)
async def create(
    body: FiscalyearCreate, _: CurrentUser, db: Annotated[AsyncSession, Depends(get_db)]
) -> RetData[FiscalyearEntity]:
    now = datetime.now(UTC)
    obj = Fiscalyear(**body.model_dump(), createdAt=now, updatedAt=now)
    db.add(obj)
    await db.flush()
    return RetData(data=FiscalyearEntity.model_validate(obj), message="Fiscalyear created")


@router.get("", response_model=RetData[list[FiscalyearEntity]])
async def find_all(_: CurrentUser, db: Annotated[AsyncSession, Depends(get_db)]) -> RetData[list[FiscalyearEntity]]:
    rows = (await db.execute(select(Fiscalyear).order_by(Fiscalyear.year.desc()))).scalars().all()
    return RetData(data=[FiscalyearEntity.model_validate(r) for r in rows], message="Fiscalyears found")


@router.get("/{fy_id}", response_model=RetData[FiscalyearEntity])
async def find_one(
    fy_id: int, _: CurrentUser, db: Annotated[AsyncSession, Depends(get_db)]
) -> RetData[FiscalyearEntity]:
    obj = await db.get(Fiscalyear, fy_id)
    if obj is None:
        return RetData(data=None, message="findOne")
    return RetData(data=FiscalyearEntity.model_validate(obj), message="Fiscalyear found")


@router.patch("/{fy_id}", response_model=RetData[FiscalyearEntity])
async def update(
    fy_id: int,
    body: FiscalyearUpdate,
    _: CurrentUser,
    db: Annotated[AsyncSession, Depends(get_db)],
) -> RetData[FiscalyearEntity]:
    obj = await db.get(Fiscalyear, fy_id)
    if obj is None:
        raise HTTPException(status_code=404, detail="Fiscalyear not found")
    for k, v in body.model_dump(exclude_none=True).items():
        setattr(obj, k, v)
    obj.updatedAt = datetime.now(UTC)
    await db.flush()
    return RetData(data=FiscalyearEntity.model_validate(obj), message="Fiscalyear updated")


@router.delete("/{fy_id}", response_model=RetData[FiscalyearEntity])
async def remove(fy_id: int, _: CurrentUser, db: Annotated[AsyncSession, Depends(get_db)]) -> RetData[FiscalyearEntity]:
    obj = await db.get(Fiscalyear, fy_id)
    if obj is None:
        raise HTTPException(status_code=404, detail="Fiscalyear not found")
    entity = FiscalyearEntity.model_validate(obj)
    return RetData(data=entity, message="Fiscalyear removed")
