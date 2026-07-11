"""``anlaesse`` router (incl. Stammblatt Excel export)."""

from __future__ import annotations

from datetime import UTC, date, datetime
from pathlib import Path
from typing import Annotated, Any

from fastapi import APIRouter, Depends, HTTPException, Query
from openpyxl import Workbook
from openpyxl.styles import Border, Font, Side
from openpyxl.worksheet.worksheet import Worksheet
from sqlalchemy import and_, func, select
from sqlalchemy.ext.asyncio import AsyncSession
from sqlalchemy.orm import selectinload

from app.core.config import get_config, load_params
from app.core.database import get_db
from app.core.logging import get_logger
from app.dependencies import CurrentUser
from app.models.adressen import Adressen
from app.models.anlaesse import Anlaesse
from app.models.meisterschaft import Meisterschaft
from app.schemas.anlaesse import AnlaesseCreate, AnlaesseEntity, AnlaesseUpdate
from app.schemas.ret_data import RetData, RetDataFile, RetDataFilePayload
from app.utils.general import (
    I_FONT_SIZE_HEADER,
    I_FONT_SIZE_ROW,
    I_FONT_SIZE_TITEL,
    set_cell_value_format,
)

logger = get_logger(__name__)

router = APIRouter(prefix="/anlaesse", tags=["Anlaesse"])

C_NAME = "C6"
C_VORNAME = "C7"
S_FIRST_ROW = 13


def _format_date_ch(d: date | datetime) -> str:
    if isinstance(d, datetime):
        d = d.date()
    return d.strftime("%d.%m.%Y")


@router.get("/overview", response_model=RetData[list[dict]])
async def get_overview(db: Annotated[AsyncSession, Depends(get_db)]) -> RetData[list[dict]]:
    params = await load_params()
    year = params.get("CLUBJAHR", "1900")
    from_d = date(int(year), 1, 1)
    to_d = date(int(year) + 1, 1, 1)
    total = await db.scalar(
        select(func.count(Anlaesse.id)).where(
            and_(Anlaesse.datum >= from_d, Anlaesse.datum < to_d, Anlaesse.nachkegeln.is_(False))
        )
    )
    today = datetime.now(UTC).date()
    upcoming = await db.scalar(
        select(func.count(Anlaesse.id)).where(
            and_(Anlaesse.datum > today, Anlaesse.datum < to_d, Anlaesse.nachkegeln.is_(False))
        )
    )
    return RetData(
        data=[
            {"label": "Total Anlässe", "value": int(total or 0)},
            {"label": "Zukünftige Anlässe", "value": int(upcoming or 0)},
        ],
        message="Übersicht Anlässe",
    )


@router.get("/getFkData", response_model=RetData[list[dict]])
async def get_fk_data(jahr: str, _: CurrentUser, db: Annotated[AsyncSession, Depends(get_db)]) -> RetData[list[dict]]:
    from_d = date(int(jahr), 1, 1)
    to_d = date(int(jahr) + 1, 1, 1)
    rows = (
        (
            await db.execute(
                select(Anlaesse)
                .where(and_(Anlaesse.datum > from_d, Anlaesse.datum < to_d))
                .order_by(Anlaesse.datum.asc())
            )
        )
        .scalars()
        .all()
    )
    return RetData(data=[{"id": r.id, "value": r.longname} for r in rows], message="getFKData")


# ---------------------------------------------------------------------------
# Stammblatt Excel export
# ---------------------------------------------------------------------------


async def _create_template(db: AsyncSession, syear: str, sheet: Worksheet, incl_points: bool) -> None:
    from_d = date(int(syear), 1, 1)
    to_d = date(int(syear) + 1, 1, 1)
    db_events = (
        (
            await db.execute(
                select(Anlaesse)
                .where(and_(Anlaesse.datum > from_d, Anlaesse.datum < to_d))
                .order_by(Anlaesse.istkegeln.desc(), Anlaesse.datum.asc(), Anlaesse.name.asc())
            )
        )
        .scalars()
        .all()
    )

    set_cell_value_format(
        sheet,
        "A2:I2",
        "CLUB/KEGELMEISTERSCHAFT",
        merge=True,
        font=Font(bold=True, size=I_FONT_SIZE_HEADER),
        align_h="center",
        align_v="center",
    )
    set_cell_value_format(
        sheet,
        "A4:I4",
        syear,
        merge=True,
        font=Font(bold=True, size=I_FONT_SIZE_HEADER),
        align_h="center",
        align_v="center",
    )
    sheet["B6"] = "Name:"
    sheet["B6"].font = Font(bold=True, size=I_FONT_SIZE_TITEL)
    sheet[C_NAME].font = Font(bold=True, size=I_FONT_SIZE_TITEL)
    sheet["B7"] = "Vorname:"
    sheet["B7"].font = Font(bold=True, size=I_FONT_SIZE_TITEL)
    sheet[C_VORNAME].font = Font(size=I_FONT_SIZE_TITEL)

    set_cell_value_format(
        sheet,
        "C11:E11",
        "Kegelmeisterschaft",
        border=True,
        merge=True,
        font=Font(bold=True, size=I_FONT_SIZE_TITEL),
    )

    row = S_FIRST_ROW - 1
    headers = [
        ("A", "Club"),
        ("B", "Datum"),
        ("H", "z Pkt."),
        ("I", "Total"),
        ("J", "Visum"),
        ("K", "eventid"),
    ]
    for col, txt in headers:
        set_cell_value_format(
            sheet,
            f"{col}{row}",
            txt,
            border=col != "K",
            font=Font(bold=True, size=I_FONT_SIZE_ROW),
        )
    set_cell_value_format(
        sheet,
        f"C{row}:G{row}",
        "Resultate",
        border=True,
        merge=True,
        font=Font(bold=True, size=I_FONT_SIZE_ROW),
    )
    sheet.column_dimensions["B"].width = 14

    club_total = 0
    for ev in db_events:
        if ev.istkegeln:
            row += 1
            value: Any = (ev.punkte or 0) if incl_points and ev.status == 1 else ""
            font = Font(size=I_FONT_SIZE_ROW, strike=ev.status != 1)
            if ev.status == 1:
                club_total += ev.punkte or 0
            set_cell_value_format(sheet, f"A{row}", value, border=True, font=font)
            set_cell_value_format(
                sheet,
                f"B{row}",
                _format_date_ch(ev.datum),
                border=True,
                font=Font(size=I_FONT_SIZE_ROW),
            )
            for c in ("C", "D", "E", "F", "G"):
                set_cell_value_format(sheet, f"{c}{row}", "", border=True, font=Font(size=I_FONT_SIZE_ROW))
            set_cell_value_format(
                sheet,
                f"H{row}",
                0 if ev.nachkegeln else 5,
                border=True,
                font=Font(size=I_FONT_SIZE_ROW),
            )
            for c in ("I", "J"):
                set_cell_value_format(sheet, f"{c}{row}", "", border=True, font=Font(size=I_FONT_SIZE_ROW))
            set_cell_value_format(sheet, f"K{row}", ev.id, font=Font(size=I_FONT_SIZE_ROW))

    row += 1
    set_cell_value_format(
        sheet,
        f"F{row}:H{row}",
        "Total Kegeln",
        border=True,
        merge=True,
        font=Font(bold=True, size=I_FONT_SIZE_ROW),
    )
    set_cell_value_format(sheet, f"I{row}", 0, border=True, font=Font(bold=True, size=I_FONT_SIZE_ROW))
    row += 2

    set_cell_value_format(
        sheet,
        f"C{row}:E{row}",
        "Clubmeisterschaft",
        border=True,
        merge=True,
        font=Font(bold=True, size=I_FONT_SIZE_TITEL),
    )

    row += 1
    for col, txt in (("A", "Club"), ("B", "Datum")):
        set_cell_value_format(sheet, f"{col}{row}", txt, border=True, font=Font(bold=True, size=I_FONT_SIZE_ROW))
    set_cell_value_format(
        sheet,
        f"C{row}:I{row}",
        "Bezeichnung",
        border=True,
        merge=True,
        font=Font(bold=True, size=I_FONT_SIZE_ROW),
    )

    for ev in db_events:
        if not ev.istkegeln:
            row += 1
            value: Any = (ev.punkte or 0) if incl_points and (ev.status or 0) > 0 else ""
            font = Font(size=I_FONT_SIZE_ROW, strike=(ev.status or 0) <= 0)
            if (ev.status or 0) > 0:
                club_total += ev.punkte or 0
            set_cell_value_format(sheet, f"A{row}", value, border=True, font=font)
            set_cell_value_format(
                sheet,
                f"B{row}",
                _format_date_ch(ev.datum),
                border=True,
                font=Font(size=I_FONT_SIZE_ROW),
            )
            set_cell_value_format(
                sheet,
                f"C{row}:I{row}",
                ev.name,
                border=True,
                merge=True,
                font=Font(size=I_FONT_SIZE_ROW),
            )
            set_cell_value_format(sheet, f"K{row}", ev.id, font=Font(size=I_FONT_SIZE_ROW))

    row += 1
    set_cell_value_format(sheet, f"B{row}", "Total Club", border=True, font=Font(bold=True, size=I_FONT_SIZE_ROW))
    set_cell_value_format(
        sheet,
        f"A{row}",
        club_total if incl_points else 0,
        border=True,
        font=Font(bold=True, size=I_FONT_SIZE_ROW),
    )

    sheet.column_dimensions["K"].hidden = True
    sheet.column_dimensions["J"].width = 17


def _fill_name(sheet: Worksheet, adresse: Adressen) -> None:
    sheet[C_NAME] = adresse.name
    sheet[C_VORNAME] = adresse.vorname


async def _fill_template(db: AsyncSession, sheet: Worksheet, adr_id: int, syear: str) -> None:
    from_d = date(int(syear), 1, 1)
    to_d = date(int(syear) + 1, 1, 1)
    rows = (
        (
            await db.execute(
                select(Meisterschaft)
                .join(Anlaesse, Anlaesse.id == Meisterschaft.eventid)
                .where(
                    and_(
                        Meisterschaft.mitgliedid == adr_id,
                        Anlaesse.datum > from_d,
                        Anlaesse.datum < to_d,
                    )
                )
                .order_by(Meisterschaft.id.asc())
            )
        )
        .scalars()
        .all()
    )
    if not rows:
        return

    club_total = 0
    kegel_total = 0
    diag_side = Side(style="thin")

    for sheet_row in sheet.iter_rows():
        for cell in sheet_row:
            if cell.column == 11 and cell.value not in (None, "", "eventid"):
                for m in rows:
                    if cell.value == m.eventid:
                        row_num = int(cell.row) if cell.row else 0
                        sheet.cell(row=row_num, column=1, value=m.punkte)
                        club_total += m.punkte or 0
                        any_wurf = any((m.wurf1, m.wurf2, m.wurf3, m.wurf4, m.wurf5))
                        if any_wurf:
                            kegel_summe = sum(
                                int(x or 0) for x in (m.wurf1, m.wurf2, m.wurf3, m.wurf4, m.wurf5, m.zusatz)
                            )
                            sheet.cell(row=row_num, column=3, value=m.wurf1)
                            sheet.cell(row=row_num, column=4, value=m.wurf2)
                            sheet.cell(row=row_num, column=5, value=m.wurf3)
                            sheet.cell(row=row_num, column=6, value=m.wurf4)
                            sheet.cell(row=row_num, column=7, value=m.wurf5)
                            sheet.cell(row=row_num, column=9, value=kegel_summe)
                            if not (m.streichresultat or False):
                                kegel_total += kegel_summe
                            else:
                                for c in sheet[row_num]:
                                    c.border = Border(
                                        top=diag_side,
                                        bottom=diag_side,
                                        left=diag_side,
                                        right=diag_side,
                                        diagonal=diag_side,
                                        diagonalUp=True,
                                        diagonalDown=True,
                                    )
                        break

    for r in range(S_FIRST_ROW, sheet.max_row + 1):
        if sheet.cell(row=r, column=6).value == "Total Kegeln":
            sheet.cell(row=r, column=9, value=kegel_total)
        elif sheet.cell(row=r, column=2).value == "Total Club":
            sheet.cell(row=r, column=1, value=club_total)


@router.get("/writestammblatt", response_model=RetDataFile)
async def write_stammblatt(
    type: int = Query(..., ge=0, le=2),
    jahr: str = Query(...),
    adresseId: int | None = None,
    _: CurrentUser = None,  # type: ignore[assignment]
    db: AsyncSession = Depends(get_db),
) -> RetDataFile:
    cfg = get_config()
    workbook = Workbook()
    if workbook.active is not None:
        workbook.remove(workbook.active)
    if type == 0:
        sheet = workbook.create_sheet("Template")
        await _create_template(db, jahr, sheet, incl_points=False)
    elif type in (1, 2):
        if adresseId:
            adr = await db.get(Adressen, int(adresseId))
            if adr is None:
                raise HTTPException(status_code=404, detail="Adresse konnte nicht gefunden werden.")
            sheet = workbook.create_sheet(adr.fullname or f"{adr.name}_{adr.vorname}")
            await _create_template(db, jahr, sheet, incl_points=type == 2)
            _fill_name(sheet, adr)
            if type == 2:
                await _fill_template(db, sheet, adr.id, jahr)
        else:
            today = datetime.now(UTC).date()
            adressen = (
                (
                    await db.execute(
                        select(Adressen)
                        .where(Adressen.austritt >= today)
                        .order_by(Adressen.name.asc(), Adressen.vorname.asc())
                    )
                )
                .scalars()
                .all()
            )
            for adr in adressen:
                sheet = workbook.create_sheet(adr.fullname or f"{adr.name}_{adr.vorname}")
                await _create_template(db, jahr, sheet, incl_points=type == 2)
                _fill_name(sheet, adr)
                if type == 2:
                    await _fill_template(db, sheet, adr.id, jahr)
    else:
        raise HTTPException(status_code=404, detail="Typ nicht gefunden.")

    filename = f"Stammblätter-{jahr}"
    if adresseId:
        filename += f"-{adresseId}"
    filename += ".xlsx"
    Path(cfg.exports).mkdir(parents=True, exist_ok=True)
    workbook.save(cfg.exports + filename)
    return RetDataFile(data=RetDataFilePayload(filename=filename), message="Excelfile erstellt")


# ---------------------------------------------------------------------------
# CRUD
# ---------------------------------------------------------------------------


@router.post("", response_model=RetData[AnlaesseEntity], status_code=201)
async def create(
    body: AnlaesseCreate, _: CurrentUser, db: Annotated[AsyncSession, Depends(get_db)]
) -> RetData[AnlaesseEntity]:
    now = datetime.now(UTC)
    longname = f"{_format_date_ch(body.datum)} - {body.name}"
    obj = Anlaesse(
        **body.model_dump(exclude={"longname"}),
        longname=longname,
        createdAt=now,
        updatedAt=now,
    )
    db.add(obj)
    await db.flush()
    await db.refresh(obj, ["anlaesse"])
    return RetData(data=AnlaesseEntity.model_validate(obj), message="Anlass erstellt")


@router.get("", response_model=RetData[list[AnlaesseEntity]])
async def find_all(
    fromJahr: int,
    toJahr: int,
    istKegeln: str | None = None,
    _: CurrentUser = None,  # type: ignore[assignment]
    db: AsyncSession = Depends(get_db),
) -> RetData[list[AnlaesseEntity]]:
    from_d = date(fromJahr, 1, 1)
    to_d = date(toJahr + 1, 1, 1)
    stmt = (
        select(Anlaesse)
        .where(and_(Anlaesse.datum > from_d, Anlaesse.datum < to_d))
        .options(selectinload(Anlaesse.anlaesse))
        .order_by(Anlaesse.datum.asc())
    )
    if istKegeln is not None:
        stmt = stmt.where(Anlaesse.istkegeln.is_(bool(istKegeln)))
    rows = (await db.execute(stmt)).scalars().all()
    return RetData(data=[AnlaesseEntity.model_validate(r) for r in rows], message="Anlässe gefunden")


@router.get("/{a_id}", response_model=RetData[AnlaesseEntity])
async def find_one(a_id: int, _: CurrentUser, db: Annotated[AsyncSession, Depends(get_db)]) -> RetData[AnlaesseEntity]:
    stmt = select(Anlaesse).where(Anlaesse.id == a_id).options(selectinload(Anlaesse.anlaesse))
    obj = (await db.execute(stmt)).scalars().first()
    if obj is None:
        return RetData(data=None, message="findOne")
    return RetData(data=AnlaesseEntity.model_validate(obj), message="Anlass gefunden")


@router.patch("/{a_id}", response_model=RetData[AnlaesseEntity])
async def update(
    a_id: int,
    body: AnlaesseUpdate,
    _: CurrentUser,
    db: Annotated[AsyncSession, Depends(get_db)],
) -> RetData[AnlaesseEntity]:
    obj = await db.get(Anlaesse, a_id)
    if obj is None:
        raise HTTPException(status_code=404, detail="Anlass konnte nicht gefunden werden")
    payload = body.model_dump(exclude_none=True)
    for k, v in payload.items():
        setattr(obj, k, v)
    obj.longname = f"{_format_date_ch(obj.datum)} - {obj.name}"
    obj.updatedAt = datetime.now(UTC)
    await db.flush()
    await db.refresh(obj, ["anlaesse"])
    return RetData(data=AnlaesseEntity.model_validate(obj), message="Anlass aktualisiert")


@router.delete("/{a_id}", response_model=RetData[AnlaesseEntity])
async def remove(a_id: int, _: CurrentUser, db: Annotated[AsyncSession, Depends(get_db)]) -> RetData[AnlaesseEntity]:
    obj = await db.get(Anlaesse, a_id)
    if obj is None:
        raise HTTPException(status_code=404, detail="Anlass konnte nicht gefunden werden")
    await db.refresh(obj, ["anlaesse"])
    entity = AnlaesseEntity.model_validate(obj)
    await db.delete(obj)
    await db.flush()
    return RetData(data=entity, message="Anlass gelöscht")
