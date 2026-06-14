"""General helpers (port of ``general.service.ts``)."""

from __future__ import annotations

from datetime import date, datetime
from typing import Any

from openpyxl.cell.cell import Cell
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.worksheet.worksheet import Worksheet

# Font sizes mirroring the NestJS constants
I_FONT_SIZE_HEADER = 18
I_FONT_SIZE_TITEL = 14
I_FONT_SIZE_ROW = 13

GERMAN_MONTHS = [
    "Januar",
    "Februar",
    "März",
    "April",
    "Mai",
    "Juni",
    "Juli",
    "August",
    "September",
    "Oktober",
    "November",
    "Dezember",
]


def format_date_long(value: date | datetime | str) -> str:
    """Return a string like ``1. Januar 2024`` (port of ``formatDateLong``)."""
    if isinstance(value, str):
        value = datetime.fromisoformat(value)
    if isinstance(value, datetime):
        value = value.date()
    return f"{value.day}. {GERMAN_MONTHS[value.month - 1]} {value.year}"


def thin_border() -> Border:
    side = Side(style="thin", color="000000")
    return Border(left=side, right=side, top=side, bottom=side)


def set_cell_value_format(
    sheet: Worksheet,
    cell_or_range: str,
    value: Any,
    border: bool = False,
    merge: bool = False,
    font: Font | None = None,
    *,
    fill: str | None = None,
    align_h: str = "left",
    align_v: str = "center",
    number_format: str | None = None,
) -> Cell:
    """Set a value (and optional merge/border/font) on a single cell or a range.

    Mirrors ``setCellValueFormat`` from amc-backend.
    """
    if merge and ":" in cell_or_range:
        sheet.merge_cells(cell_or_range)
        first_cell = cell_or_range.split(":")[0]
    else:
        first_cell = cell_or_range
    cell: Cell = sheet[first_cell]
    cell.value = value
    if font is not None:
        cell.font = font
    if border:
        cell.border = thin_border()
    if fill:
        cell.fill = PatternFill(fill_type="solid", fgColor=fill)
    cell.alignment = Alignment(horizontal=align_h, vertical=align_v, wrap_text=False)
    if number_format:
        cell.number_format = number_format
    return cell


def font_header(bold: bool = True) -> Font:
    return Font(name="Calibri", size=I_FONT_SIZE_HEADER, bold=bold)


def font_title(bold: bool = True) -> Font:
    return Font(name="Calibri", size=I_FONT_SIZE_TITEL, bold=bold)


def font_row(bold: bool = False) -> Font:
    return Font(name="Calibri", size=I_FONT_SIZE_ROW, bold=bold)
